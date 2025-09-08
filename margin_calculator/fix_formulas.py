from openpyxl import load_workbook

def fix_excel_formulas():
    wb = load_workbook('/mnt/c/Users/redsk/OneDrive/デスクトップ/mikael_project/margin_calculator/마진_분석_계산기.xlsx')
    main_sheet = wb['메인_상품별_마진_분석']
    
    print("🔧 수식 수정 중...")
    
    # 수식 수정이 필요한 행들 (2~20행)
    for row in range(2, 21):
        # M열(13열): 플랫폼 수수료(금액) = E * (H/100)
        main_sheet.cell(row=row, column=13).value = f"=E{row}*(H{row}/100)"
        
        # N열(14열): 정산 예정 금액 = E - M
        main_sheet.cell(row=row, column=14).value = f"=E{row}-M{row}"
        
        # O열(15열): 총 비용 = IF(판매방식="직접 배송", 직접배송비용, 풀필먼트비용)
        main_sheet.cell(row=row, column=15).value = f'=IF(B{row}="직접 배송",C{row}+I{row}+J{row}+L{row},C{row}+K{row}+L{row})'
        
        # P열(16열): 최종 수익 = N - O  
        main_sheet.cell(row=row, column=16).value = f"=N{row}-O{row}"
        
        # Q열(17열): 마진율 = P / E
        cell_q = main_sheet.cell(row=row, column=17)
        cell_q.value = f"=IFERROR(P{row}/E{row},0)"
        cell_q.number_format = '0.0%'
        
        # R열(18열): 납부 예상 부가세
        main_sheet.cell(row=row, column=18).value = f'=(IF(F{row}="포함",E{row}/11,E{row}*0.1))-(IF(D{row}="포함",C{row}/11,C{row}*0.1))'
    
    # 새 파일로 저장
    wb.save('/mnt/c/Users/redsk/OneDrive/デスクトップ/mikael_project/margin_calculator/마진_분석_계산기_수정.xlsx')
    wb.close()
    
    print("✅ 수식 수정 완료!")

if __name__ == "__main__":
    fix_excel_formulas()