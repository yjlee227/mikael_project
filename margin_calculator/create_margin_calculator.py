from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_margin_calculator():
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create main sheet
    main_sheet = wb.create_sheet("메인_상품별_마진_분석")
    
    # Create settings sheet
    settings_sheet = wb.create_sheet("설정_각종_수수료")
    
    # Setup main sheet headers  
    main_headers = [
        "상품명", "판매 방식", "매입가", "매입가 부가세", "판매가", "판매가 부가세",
        "플랫폼", "플랫폼 수수료 (%)", "배송비", "포장비", "풀필먼트 수수료", 
        "기타 비용", "플랫폼 수수료(금액)", "정산 예정 금액", "총 비용", 
        "최종 수익 (마진)", "마진율 (%)", "납부 예상 부가세"
    ]
    
    # Add headers to main sheet
    for col_idx, header in enumerate(main_headers, 1):
        cell = main_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Setup settings sheet
    settings_headers = ["플랫폼명", "수수료율 (%)", "기본 풀필먼트 수수료"]
    
    # Add headers to settings sheet
    for col_idx, header in enumerate(settings_headers, 1):
        cell = settings_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample platform data with fulfillment info
    platforms_data = [
        ["네이버 스마트스토어", 5.0, 2500],
        ["쿠팡", 8.0, 3000],
        ["11번가", 6.0, 2000],
        ["G마켓", 7.0, 2200],
        ["옥션", 7.0, 2200],
        ["티몬", 15.0, 2800],
        ["위메프", 15.0, 2800],
        ["인터파크", 6.0, 2100]
    ]
    
    for row_idx, (platform, fee, fulfillment) in enumerate(platforms_data, 2):
        settings_sheet.cell(row=row_idx, column=1).value = platform
        settings_sheet.cell(row=row_idx, column=2).value = fee
        settings_sheet.cell(row=row_idx, column=3).value = fulfillment
    
    # Add formulas to main sheet (row 2 as example)
    row = 2
    
    # H열: 플랫폼 수수료 (%) - VLOOKUP
    main_sheet.cell(row=row, column=8).value = f"=VLOOKUP(G{row},설정_각종_수수료!A:B,2,FALSE)"
    
    # M열: 플랫폼 수수료(금액)
    main_sheet.cell(row=row, column=13).value = f"=E{row}*(H{row}/100)"
    
    # N열: 정산 예정 금액
    main_sheet.cell(row=row, column=14).value = f"=E{row}-M{row}"
    
    # O열: 총 비용 (판매 방식에 따라)
    main_sheet.cell(row=row, column=15).value = f'=IF(B{row}="직접 배송",C{row}+I{row}+J{row}+L{row},C{row}+K{row}+L{row})'
    
    # P열: 최종 수익 (마진)
    main_sheet.cell(row=row, column=16).value = f"=N{row}-O{row}"
    
    # Q열: 마진율 (%)
    cell_q = main_sheet.cell(row=row, column=17)
    cell_q.value = f"=IFERROR(P{row}/E{row},0)"
    cell_q.number_format = '0.0%'
    
    # K열: 풀필먼트 수수료 - 플랫폼 선택 시 자동으로 기본값 적용
    main_sheet.cell(row=row, column=11).value = f"=IF(G{row}=\"\",\"\",VLOOKUP(G{row},설정_각종_수수료!A:C,3,FALSE))"
    
    # R열: 납부 예상 부가세
    main_sheet.cell(row=row, column=18).value = f'=(IF(F{row}="포함",E{row}/11,E{row}*0.1))-(IF(D{row}="포함",C{row}/11,C{row}*0.1))'
    
    # Copy formulas to additional rows (3-20)
    for target_row in range(3, 21):
        # Copy and adjust formulas
        for col in [8, 11, 13, 14, 15, 16, 17, 18]:
            source_cell = main_sheet.cell(row=2, column=col)
            target_cell = main_sheet.cell(row=target_row, column=col)
            
            # Replace row numbers in formula
            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                formula = source_cell.value.replace(f'{2}', f'{target_row}')
                target_cell.value = formula
                
                # Apply number format for percentage column
                if col == 17:
                    target_cell.number_format = '0.0%'
    
    # Add data validation for dropdowns
    
    # B열: 판매 방식 dropdown
    sales_method_dv = DataValidation(type="list", formula1='"직접 배송,풀필먼트"')
    sales_method_dv.add(f'B2:B1000')
    main_sheet.add_data_validation(sales_method_dv)
    
    # D열: 매입가 부가세 dropdown  
    purchase_vat_dv = DataValidation(type="list", formula1='"포함,미포함"')
    purchase_vat_dv.add(f'D2:D1000')
    main_sheet.add_data_validation(purchase_vat_dv)
    
    # F열: 판매가 부가세 dropdown
    sale_vat_dv = DataValidation(type="list", formula1='"포함,미포함"')
    sale_vat_dv.add(f'F2:F1000')
    main_sheet.add_data_validation(sale_vat_dv)
    
    # G열: 플랫폼 dropdown - settings sheet 참조
    platform_dv = DataValidation(type="list", formula1=f"설정_각종_수수료!$A$2:$A${len(platforms_data)+1}")
    platform_dv.add(f'G2:G1000')
    main_sheet.add_data_validation(platform_dv)
    
    # Auto-adjust column widths
    for sheet in [main_sheet, settings_sheet]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    wb.save("/mnt/c/Users/redsk/OneDrive/デスクトップ/mikael_project/margin_calculator/마진_분석_계산기.xlsx")
    print("Excel file created successfully: 마진_분석_계산기.xlsx")

if __name__ == "__main__":
    create_margin_calculator()