from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment

def add_conditional_blocking():
    wb = load_workbook('/mnt/c/Users/redsk/OneDrive/デスクトップ/mikael_project/margin_calculator/마진_분석_계산기_수정.xlsx')
    main_sheet = wb['메인_상품별_마진_분석']
    
    print("🚫 조건부 입력 차단 기능 추가 중...")
    
    # 1. 배송비(I열) - 풀필먼트일 때 입력 차단
    shipping_dv = DataValidation(
        type="custom", 
        formula1='B2<>"풀필먼트"',
        errorTitle="입력 제한",
        error="풀필먼트 방식에서는 배송비를 입력할 수 없습니다."
    )
    shipping_dv.add('I2:I1000')
    main_sheet.add_data_validation(shipping_dv)
    
    # 2. 포장비(J열) - 풀필먼트일 때 입력 차단  
    packaging_dv = DataValidation(
        type="custom",
        formula1='B2<>"풀필먼트"',
        errorTitle="입력 제한", 
        error="풀필먼트 방식에서는 포장비를 입력할 수 없습니다."
    )
    packaging_dv.add('J2:J1000')
    main_sheet.add_data_validation(packaging_dv)
    
    # 3. 풀필먼트 수수료(K열) - 직접 배송일 때 수정 차단 (자동입력이므로)
    # K열은 수식으로 자동 입력되므로 데이터 검증 대신 조건부 서식만 적용
    
    # 4. 조건부 서식 - 비활성화된 셀 시각적 표시
    
    # 배송비 열 - 풀필먼트일 때 회색 배경
    gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # 각 행별로 조건부 서식 적용
    for row in range(2, 1001):
        # 배송비 - 해당 행의 B열이 "풀필먼트"일 때
        shipping_rule = FormulaRule(formula=[f'$B{row}="풀필먼트"'], fill=gray_fill)
        main_sheet.conditional_formatting.add(f'I{row}', shipping_rule)
        
        # 포장비 - 해당 행의 B열이 "풀필먼트"일 때  
        packaging_rule = FormulaRule(formula=[f'$B{row}="풀필먼트"'], fill=gray_fill)
        main_sheet.conditional_formatting.add(f'J{row}', packaging_rule)
        
        # 풀필먼트 수수료 - 해당 행의 B열이 "직접 배송"일 때
        fulfillment_rule = FormulaRule(formula=[f'$B{row}="직접 배송"'], fill=gray_fill)
        main_sheet.conditional_formatting.add(f'K{row}', fulfillment_rule)
    
    # 5. 헤더에 도움말 추가
    main_sheet['I1'].value = "배송비\n(직접배송만)"
    main_sheet['J1'].value = "포장비\n(직접배송만)"  
    main_sheet['K1'].value = "풀필먼트 수수료\n(자동입력)"
    
    # 헤더 스타일 조정
    for col in ['I1', 'J1', 'K1']:
        cell = main_sheet[col]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 파일 저장
    wb.save('/mnt/c/Users/redsk/OneDrive/デスクトップ/mikael_project/margin_calculator/마진_분석_계산기_최종.xlsx')
    wb.close()
    
    print("✅ 조건부 입력 차단 기능 추가 완료!")
    print("📁 새 파일: 마진_분석_계산기_최종.xlsx")

if __name__ == "__main__":
    add_conditional_blocking()