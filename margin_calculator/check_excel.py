from openpyxl import load_workbook

def check_excel_file():
    try:
        # Excel 파일 열기
        wb = load_workbook('/mnt/c/Users/redsk/OneDrive/デスクトップ/mikael_project/margin_calculator/마진_분석_계산기.xlsx')
        
        print("📊 Excel 파일 정보:")
        print(f"워크시트 목록: {wb.sheetnames}")
        
        # 메인 시트 확인
        main_sheet = wb['메인_상품별_마진_분석']
        print(f"\n📋 메인 시트 정보:")
        print(f"최대 행: {main_sheet.max_row}")
        print(f"최대 열: {main_sheet.max_column}")
        
        # 헤더 확인
        print(f"\n📝 헤더 정보:")
        headers = []
        for col in range(1, main_sheet.max_column + 1):
            cell_value = main_sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(f"{col}열: {cell_value}")
        
        for header in headers:
            print(header)
        
        # 설정 시트 확인
        if '설정_각종_수수료' in wb.sheetnames:
            settings_sheet = wb['설정_각종_수수료']
            print(f"\n⚙️ 설정 시트 정보:")
            print(f"최대 행: {settings_sheet.max_row}")
            
            print("\n🏪 플랫폼 데이터:")
            for row in range(2, settings_sheet.max_row + 1):
                platform = settings_sheet.cell(row=row, column=1).value
                fee = settings_sheet.cell(row=row, column=2).value
                fulfillment = settings_sheet.cell(row=row, column=3).value
                if platform:
                    print(f"{platform}: {fee}%, 풀필먼트 {fulfillment}원")
        
        # 수식 확인 (2행)
        print(f"\n🔢 수식 확인 (2행):")
        formulas = []
        for col in range(1, main_sheet.max_column + 1):
            cell = main_sheet.cell(row=2, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append(f"{col}열: {cell.value}")
        
        for formula in formulas:
            print(formula)
            
        # 데이터 검증 확인
        print(f"\n✅ 데이터 검증 규칙:")
        for dv in main_sheet.data_validations.dataValidation:
            print(f"범위: {dv.sqref}, 타입: {dv.type}, 공식: {dv.formula1}")
        
        wb.close()
        print(f"\n✅ Excel 파일이 정상적으로 생성되었습니다!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")

if __name__ == "__main__":
    check_excel_file()